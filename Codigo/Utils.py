import openpyxl.workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep
from pathlib import Path
from threading import Thread, Event
import os
import openpyxl
import re

qr_code_verificado = False
msgs_carregadas = False

class Funcao_Botoes_Pesquisar:
    def __init__(self, tela_principal, caminho_planilha, caminho_img):
        from tkinter import filedialog

        self.tela_principal = tela_principal
        self.caminho_planilha = caminho_planilha
        self.caminho_img = caminho_img

        self.filedialog = filedialog

    def abrir_explorador_planilha(self):
        from Tela_Auxiliar import Tela_Auxiliar

        if self.tela_principal.botao_procurar_planilha._state == 'disable':
            return

        extensao_valida = ['.xlsx']

        self.pesquisar_caminho_planilha = self.filedialog.askopenfilename(
            title='Pesquisar',
            filetypes=(('Planilhas Excel', '*.xlsx'), ('Todos os Arquivos', '*.*'))
        )

        if Path(self.pesquisar_caminho_planilha).suffix in extensao_valida:
            self.caminho_planilha.configure(state='normal')

            self.caminho_planilha.delete(0, 'end')
            self.caminho_planilha.insert(0, f'{self.pesquisar_caminho_planilha}')

            self.caminho_planilha.configure(state='disable')
        else:
            Tela_Auxiliar(self.tela_principal, texto='Extensão do arquivo inválida!\nPor favor, selecione uma planilha com as extensão: ".xlsx"')

    def abrir_explorador_img(self):
        from Tela_Auxiliar import Tela_Auxiliar

        if self.tela_principal.botao_procurar_img._state == 'disable':
            return

        extensoes_validas = ['.png','.jpg','.jpeg']

        self.pesquisar_caminho_img = self.filedialog.askopenfilename(
            title='Pesquisar',
            filetypes=(('Arquivos de Imagem', '*.png *.jpg *.jpeg'), ('Todos os Arquivos', '*.*')),
        )

        if Path(self.pesquisar_caminho_img).suffix in extensoes_validas:
            self.caminho_img.configure(state='normal')

            self.caminho_img.delete(0, 'end')
            self.caminho_img.insert(0, f'{self.pesquisar_caminho_img}')

            self.caminho_img.configure(state='disable')
            self.tela_principal.opcao_img.select()
        else:
            Tela_Auxiliar(self.tela_principal, texto='Extensão do arquivo inválida!\nPor favor, selecione uma imagem com as extensões:\n".png", ".jpg" ou ".jpeg"')


class Funcao_Botoes_Acoes:
    def __init__(self, tela_principal, caminho_planilha, caminho_img):
        self.tela_principal = tela_principal
        self.caminho_planilha = caminho_planilha
        self.caminho_img = caminho_img

        self.navegador = None
        self.interromper = False

        self.navegador_pronto = Event()
        
    def botao_iniciar_envio(self):
        from Tela_Auxiliar import Tela_Auxiliar

        if hasattr(self, 'navegador') and self.navegador:
            try:
                self.navegador.encerrar_navegador()
            except Exception as e:
                self.tela_principal.texto_area_log(texto=f'Erro ao encerrar navegador anterior: {str(e)}')

        self.opcao_img = self.tela_principal.checar.get()
        self.planilha_selecionada = self.caminho_planilha.get()
        self.img_selecionada = self.caminho_img.get()
        self.campo_msg_preenchido = self.tela_principal.mensagem_digitada.get('1.0', 'end').strip()

        if self.tela_principal.botao_iniciar._state == 'disable':
            return

        if not self.planilha_selecionada or self.campo_msg_preenchido == '':
            Tela_Auxiliar(self.tela_principal, texto='O caminho da planilha e o campo da mensagem\nnão podem estar vazios!')
        else:
            if self.opcao_img == 'on':
                if self.img_selecionada:
                    Tela_Auxiliar(self.tela_principal, True, self.execucao_geral, 'Realmente deseja prosseguir com o envio das mensagens?\n\nCertifique-se de que a planilha está no formato adequado.')
                else:
                    Tela_Auxiliar(self.tela_principal, texto='Campo "Enviar mensagem com a imagem" está marcado,\nmas o caminho da imagem está vazio!')
            else:
                Tela_Auxiliar(self.tela_principal, True, self.execucao_geral, 'Realmente deseja prosseguir com o envio das mensagens\nsem a imagem?\n\nCertifique-se de que a planilha está no formato adequado.')

    def botao_interromper(self):
        if self.tela_principal.botao_interromper._state == 'disable':
            return
        
        self.tela_principal.botao_interromper.configure(state='disable')

        self.tela_principal.texto_area_log(texto='Interrompendo envio de mensagens...')

        try:
            self.navegador.encerrar_navegador()
            self.navegador = None
        except Exception as e:
            self.tela_principal.texto_area_log(texto=f'Erro ao encerrar navegador: {str(e)}')
            if self.navegador:
                self.navegador = None
        
        self.interromper = True
        
    def iniciar_navegador(self):
        try:
            self.tela_principal.texto_area_log(texto='Iniciando navegador...')
            self.navegador = Navegador(self.tela_principal)
            self.navegador_pronto.set()
        except Exception as e:
            self.tela_principal.texto_area_log(texto=f'Erro ao iniciar navegador: {str(e)}')
            self.restaurar_interface()
            if self.navegador:
                self.navegador = None
        
    def enviar_msgs_navegador(self):
        import urllib
        from urllib3.exceptions import NewConnectionError

        try:
            self.navegador_pronto.wait()

            planilha_clientes = openpyxl.load_workbook(self.planilha_selecionada)
            aba_ativa = planilha_clientes.active

            self.tela_principal.texto_area_log(texto='Iniciando envio de mensagens...')
            self.tela_principal.botao_interromper.configure(state='normal')

            for linha in aba_ativa.iter_rows(min_row=2):
                if all(linha[celula].value is None for celula in range(0, 4)):
                    continue

                if self.interromper:
                    self.tela_principal.texto_area_log(texto='Envio interrompido pelo usuário!')
                    if self.navegador:
                        try:
                            self.navegador.encerrar_navegador()
                            self.navegador = None
                        except Exception as e:
                            self.tela_principal.texto_area_log(texto=f'Erro ao encerrar navegador: {str(e)}')

                    self.salvar_planilha(planilha_clientes, self.planilha_selecionada)
                    break

                try:
                    nome_contato = linha[0].value.strip()
                    telefone = str(linha[1].value).strip()
                    nome_vendedor = linha[2].value.strip()

                    tel_analise = re.sub(r'\D', '', telefone)
                    tel_letra = re.search(r'[a-zA-Z]', telefone)

                    if not nome_contato or not telefone or not nome_vendedor:
                        linha[3].value = 'Erro: Dados incompletos'
                        self.salvar_planilha(planilha_clientes, self.planilha_selecionada)
                        continue

                    if len(tel_analise) < 11 or tel_letra:
                        linha[3].value = 'Erro: número de telefone inválido'
                        self.salvar_planilha(planilha_clientes, self.planilha_selecionada)
                        continue

                    if linha[3].value in ['Erro em enviar a mensagem','Mensagem enviada com sucesso','Erro: Dados incompletos']:
                        continue

                    mensagem_modificada = self.campo_msg_preenchido.replace('CONTATO', f'{nome_contato}')
                    mensagem_modificada = mensagem_modificada.replace('VENDEDOR', f'{nome_vendedor}')
                    mensagem_modificada = urllib.parse.quote(mensagem_modificada)

                    retorno_status = self.navegador.script_envio_msgs(
                        telefone,
                        mensagem_modificada,
                        self.img_selecionada if self.opcao_img == 'on' else None
                    )

                    linha[3].value = retorno_status
                
                except NewConnectionError:
                    pass

                finally:
                    self.salvar_planilha(planilha_clientes, self.planilha_selecionada)

        except:
            pass

        finally:
            if self.navegador:
                self.navegador.encerrar_navegador()
                self.navegador = None

            self.restaurar_interface()
            self.tela_principal.texto_area_log(texto='Envio de mensagens finalizado!')
            self.tela_principal.botao_interromper.configure(state='disable')

    def salvar_planilha(self, planilha, caminho_original):
        caminho_temporario = caminho_original + '.temp'

        try:
            planilha.save(caminho_temporario)
            os.replace(caminho_temporario, caminho_original)
        except:
            self.tela_principal.texto_area_log(texto=f'Erro ao salvar a planilha!')

            if os.path.exists(caminho_temporario):
                os.remove(caminho_temporario)

    def restaurar_interface(self):
        self.tela_principal.botao_iniciar.configure(state='normal')
        self.tela_principal.mensagem_digitada.configure(state='normal')
        self.tela_principal.opcao_img.configure(state='normal')
        self.tela_principal.botao_procurar_planilha.configure(state='normal')
        self.tela_principal.botao_procurar_img.configure(state='normal')

    def verificar_planilha_aberta(self):
        try:
            planilha = openpyxl.load_workbook(self.planilha_selecionada)
            planilha.save(self.planilha_selecionada)
            return False
        except:
            return True
            
    def execucao_geral(self):
        if self.verificar_planilha_aberta():
            from Tela_Auxiliar import Tela_Auxiliar

            Tela_Auxiliar(self.tela_principal, texto='A planilha selecionada está aberta!\nPor favor, feche-a e tente novamente!')
            return

        self.interromper = False
        self.navegador_pronto.clear()

        self.tela_principal.botao_iniciar.configure(state='disable')
        self.tela_principal.mensagem_digitada.configure(state='disable')
        self.tela_principal.opcao_img.configure(state='disable')
        self.tela_principal.botao_procurar_planilha.configure(state='disable')
        self.tela_principal.botao_procurar_img.configure(state='disable')
        
        Thread(target=self.iniciar_navegador, daemon=True).start()

        Thread(target=self.enviar_msgs_navegador, daemon=True).start()
      

class Navegador:
    def __init__(self, tela_principal):
        self.tela_principal = tela_principal

        self.navegador = None
        self.wait = None

        try:
            self.iniciar_navegador()
        except Exception as e:
            self.tela_principal.texto_area_log(texto=f'Erro ao iniciar navegador: {str(e)}')
            if self.navegador:
                self.encerrar_navegador()
            raise e

    def config_navegador(self):
        altura, largura = self.calcular_tam_tela_navegador()

        options = webdriver.ChromeOptions()

        perfil_chrome = self.buscar_usuario_atual()
        
        argumentos = [
            f'user-data-dir={perfil_chrome}',
            '--block-new-web-contents',
            '--disable-notifications',
            '--no-default-browser-check',
            '--disable-features=ExternalProtocolDialog',
            '--lang=pt-BR',
            '--window-position=0,0',
            f'--window-size={largura},{altura}'
        ]

        for argumento in argumentos:
            options.add_argument(argumento)

        options.add_experimental_option("detach", True)
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)


        navegador = webdriver.Chrome(options=options)

        wait = WebDriverWait(navegador, timeout=10)

        return navegador, wait

    def calcular_tam_tela_navegador(self):
        from screeninfo import get_monitors
        
        monitor = get_monitors()[0]

        altura = monitor.height
        largura = monitor.width

        return altura, round(largura/2)

    def buscar_usuario_atual(self):
        pasta_usuario_atual = os.environ["USERPROFILE"]
    
        perfil_chrome = f"{pasta_usuario_atual}\\AppData\\Local\\Google\\Chrome\\User Data\\Default"

        return perfil_chrome

    def verificar_QR_code(self):
        global qr_code_verificado

        if qr_code_verificado:
            sleep(1)
            return

        self.tela_principal.texto_area_log(texto='Verificando QR code...')

        while True:
            try:
                qr_code = self.wait.until(EC.presence_of_element_located((By.XPATH, '//canvas[@aria-label="Scan this QR code to link a device!"]')))
                if qr_code:
                    sleep(2)
                    
            except:
                self.tela_principal.texto_area_log(texto='QR code verificado!')
                qr_code_verificado = True
                sleep(1)
                break

    def aguardar_carregamento_msgs(self):
        global msgs_carregadas

        self.tela_principal.texto_area_log(texto='Aguardando carregamento de mensagens...')
        while True:
            try:
                carre_msgs = self.navegador.find_element(by=By.XPATH, value='//div[@class="x1c3i2sq x14ug900 xk82a7y x1sy10c2"]')
                if carre_msgs:
                    sleep(2)
                else:
                    self.tela_principal.texto_area_log(texto='Mensagens carregadas!')
                    msgs_carregadas = True
                    break
            except:
                self.tela_principal.texto_area_log(texto='Mensagens carregadas!')
                msgs_carregadas = True
                break

    def iniciar_navegador(self):
        global msgs_carregadas

        if self.navegador:
            self.encerrar_navegador()

        self.navegador, self.wait = self.config_navegador()
        sleep(1)
        self.navegador.get('https://web.whatsapp.com/')
        sleep(5)
        self.verificar_QR_code()

        if not msgs_carregadas:
            self.aguardar_carregamento_msgs()

    def encerrar_navegador(self):
        try:
            if self.navegador:
                self.navegador.quit()
        finally:
            self.navegador = None
            self.wait = None

    def script_envio_msgs(self, telefone, mensagem, imagem):
        self.navegador.get(f'https://web.whatsapp.com/send?phone={telefone}&text={mensagem}')
        sleep(3)

        try:           
            if imagem:
                botao_mais = self.wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div/div[3]/div/div[4]/div/footer/div[1]/div/span/div/div[1]/div/button/span')))
                botao_mais.click()
                sleep(2)

                botao_fotos = self.wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div/div/span[5]/div/ul/div/div/div[2]/li/div/input')))
                botao_fotos.send_keys(imagem)
                sleep(5)

                enter = self.wait.until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/div/div[2]/div[2]/span/div/div/div/div[2]/div/div[2]/div[2]/div/div')))
                enter.click()
                sleep(6)
            else:
                sleep(3)
                enter = self.wait.until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[3]/div/div[4]/div/footer/div[1]/div/span/div/div[2]/div[2]/button/span')))
                enter.click()
                sleep(6)

            status = 'Mensagem enviada com sucesso'
        except:
            status = 'Erro em enviar a mensagem'
            
        return status